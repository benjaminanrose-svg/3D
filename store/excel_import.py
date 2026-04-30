"""
Importación y exportación de productos desde/hacia Excel para Gflex3D.

Columnas de la plantilla:
  nombre | categoria | precio | descripcion_corta | descripcion |
  materiales | vehiculos_compatibles | dias_fabricacion | stock |
  precio_minimo | destacado | estado
"""
import io
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from django.http import HttpResponse

from .models import Product, Category, Material, ProductVariant


# ── COLUMNAS ─────────────────────────────────────────────────────────────────

COLUMNS = [
    ('nombre',            'Nombre del producto *',           28),
    ('categoria',         'Categoría *',                     20),
    ('precio',            'Precio base CLP *',               14),
    ('descripcion_corta', 'Descripción corta',               32),
    ('descripcion',       'Descripción completa',            42),
    ('materiales',        'Materiales (separados por /)',    26),
    ('vehiculos',         'Vehículos compatibles',           30),
    ('dias_fabricacion',  'Días de fabricación',             14),
    ('destacado',         'Destacado',                       12),
    ('estado',            'Estado',                          16),
]

HEADERS    = [c[1] for c in COLUMNS]
COL_WIDTHS = [c[2] for c in COLUMNS]
COL_KEYS   = [c[0] for c in COLUMNS]

# Índices de columna (1-based) para validaciones
COL_CATEGORIA    = 2
COL_PRECIO       = 3
COL_DIAS         = 8
COL_DESTACADO    = 9
COL_ESTADO       = 10

DATA_START_ROW = 4
DATA_END_ROW   = 203  # 200 filas de datos


# ── ESTILOS ───────────────────────────────────────────────────────────────────

def _header_styles():
    border_side = Side(style='thin', color='333333')
    return {
        'header_fill':  PatternFill('solid', fgColor='0A0A0B'),
        'header_font':  Font(name='Calibri', bold=True, color='FF5C00', size=11),
        'subhead_fill': PatternFill('solid', fgColor='1A1C20'),
        'subhead_font': Font(name='Calibri', color='F5F3EF', size=10),
        'example_fill': PatternFill('solid', fgColor='1A1C20'),
        'example_font': Font(name='Calibri', color='C5C7CC', size=10),
        'empty_fill':   PatternFill('solid', fgColor='111214'),
        'empty_font':   Font(name='Calibri', color='555960', size=10),
        'cell_border':  Border(
            left=border_side, right=border_side,
            bottom=border_side, top=border_side
        ),
        'required_fill': PatternFill('solid', fgColor='1C0A00'),
    }


# ── GENERAR PLANTILLA ─────────────────────────────────────────────────────────

def generate_template(categories, materials):
    """Genera el archivo Excel plantilla con dropdowns y validación de datos."""
    wb  = openpyxl.Workbook()
    s   = _header_styles()

    # ── Hoja de referencias (primero, para los rangos con nombre) ─────────────
    ws2 = wb.active
    ws2.title = 'Referencias'
    _build_reference_sheet(ws2, categories, materials)

    # ── Hoja principal ────────────────────────────────────────────────────────
    ws = wb.create_sheet('Productos', 0)
    wb.active = ws

    _build_header(ws, s)
    _build_example_rows(ws, s)
    _build_empty_rows(ws, s)
    _add_data_validations(wb, ws, categories, materials)

    ws.freeze_panes = 'A4'
    ws.sheet_view.showGridLines = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_header(ws, s):
    total_cols = len(COLUMNS)

    # Fila 1 — título
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    title_cell            = ws['A1']
    title_cell.value      = 'GFLEX3D — Importación de productos'
    title_cell.font       = Font(name='Calibri', bold=True, color='FF5C00', size=14)
    title_cell.fill       = PatternFill('solid', fgColor='0A0A0B')
    title_cell.alignment  = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2 — instrucción
    ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    inst            = ws['A2']
    inst.value      = ('Completa desde la fila 4.  '
                       'Columnas con * son obligatorias.  '
                       'Las celdas amarillas tienen menú desplegable — solo haz clic en la celda.')
    inst.font       = Font(name='Calibri', color='8A8F9A', size=10, italic=True)
    inst.fill       = PatternFill('solid', fgColor='111214')
    inst.alignment  = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # Fila 3 — encabezados
    required_cols = {COL_CATEGORIA, COL_PRECIO, 1}  # nombre, categoria, precio
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell           = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = s['header_font']
        cell.fill      = s['header_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = s['cell_border']
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 32


def _build_example_rows(ws, s):
    examples = [
        ['Buje Estabilizador Delantero x2', 'Bujes', 24900,
         'Buje de poliuretano para estabilizador',
         'Fabricado en poliuretano de alta durabilidad.',
         'Poliuretano 80A', 'Toyota Corolla\nNissan Sentra',
         7, 'si', 'activo'],
        ['Manguera Radiador Superior Universal', 'Mangueras', 12900,
         'Manguera de silicona para radiador',
         'Alta resistencia al calor.',
         'Caucho EPDM', 'Toyota\nNissan\nHyundai',
         5, 'no', 'activo'],
        ['Guardapolvos Amortiguador Trasero', 'Guardapolvos', 8500,
         'Guardapolvos en caucho NBR',
         'Protege el amortiguador del polvo y la humedad.',
         'Caucho NBR', 'Suzuki Swift\nSuzuki Baleno',
         7, 'no', 'pedido'],
    ]

    dropdown_cols = {COL_CATEGORIA, COL_DESTACADO, COL_ESTADO}
    highlight_fill = PatternFill('solid', fgColor='1F1200')

    for row_idx, row_data in enumerate(examples, start=DATA_START_ROW):
        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = s['example_font']
            cell.fill      = highlight_fill if col_idx in dropdown_cols else s['example_fill']
            cell.border    = s['cell_border']
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[row_idx].height = 36


def _build_empty_rows(ws, s):
    dropdown_cols = {COL_CATEGORIA, COL_DESTACADO, COL_ESTADO}
    dropdown_fill = PatternFill('solid', fgColor='1A1500')

    for row_idx in range(DATA_START_ROW + 3, DATA_END_ROW + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell        = ws.cell(row=row_idx, column=col_idx, value='')
            cell.font   = s['empty_font']
            cell.fill   = dropdown_fill if col_idx in dropdown_cols else s['empty_fill']
            cell.border = s['cell_border']
        ws.row_dimensions[row_idx].height = 20


def _add_data_validations(wb, ws, categories, materials):
    """Agrega dropdowns y validaciones a las columnas clave."""

    col_cat   = get_column_letter(COL_CATEGORIA)
    col_dest  = get_column_letter(COL_DESTACADO)
    col_est   = get_column_letter(COL_ESTADO)
    col_precio = get_column_letter(COL_PRECIO)
    col_dias  = get_column_letter(COL_DIAS)

    # ── Dropdown: Estado ──────────────────────────────────────────────────────
    dv_estado = DataValidation(
        type='list',
        formula1='"activo,inactivo,pedido"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle='Valor inválido',
        error='Usa: activo, inactivo o pedido',
        showInputMessage=True,
        promptTitle='Estado del producto',
        prompt='Selecciona: activo, inactivo o pedido',
    )
    ws.add_data_validation(dv_estado)
    dv_estado.sqref = f'{col_est}{DATA_START_ROW}:{col_est}{DATA_END_ROW}'

    # ── Dropdown: Destacado ───────────────────────────────────────────────────
    dv_destacado = DataValidation(
        type='list',
        formula1='"si,no"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle='Valor inválido',
        error='Escribe si o no',
        showInputMessage=True,
        promptTitle='¿Producto destacado?',
        prompt='Selecciona si o no',
    )
    ws.add_data_validation(dv_destacado)
    dv_destacado.sqref = f'{col_dest}{DATA_START_ROW}:{col_dest}{DATA_END_ROW}'

    # ── Dropdown: Categoría (desde hoja Referencias) ──────────────────────────
    cat_list = [c.name for c in categories]
    if cat_list:
        # Registra un rango con nombre para que el dropdown funcione cross-sheet
        max_cat_row = 3 + len(cat_list)
        wb.defined_names['_Categorias'] = DefinedName(
            '_Categorias',
            attr_text=f'Referencias!$A$4:$A${max_cat_row}'
        )
        dv_cat = DataValidation(
            type='list',
            formula1='=_Categorias',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
            showInputMessage=True,
            promptTitle='Categoría',
            prompt='Elige una categoría de la lista o escribe una nueva (se creará automáticamente)',
        )
    else:
        dv_cat = DataValidation(
            type='list',
            formula1='"Sin categorías aún"',
            allow_blank=True,
            showDropDown=False,
            showInputMessage=True,
            promptTitle='Categoría',
            prompt='Escribe el nombre de la categoría (se creará si no existe)',
        )
    ws.add_data_validation(dv_cat)
    dv_cat.sqref = f'{col_cat}{DATA_START_ROW}:{col_cat}{DATA_END_ROW}'

    # ── Validación: Precio (número positivo) ──────────────────────────────────
    dv_precio = DataValidation(
        type='whole',
        operator='greaterThan',
        formula1='0',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='Precio inválido',
        error='El precio debe ser un número entero mayor que 0 (en pesos CLP)',
        showInputMessage=True,
        promptTitle='Precio en CLP',
        prompt='Ingresa el precio en pesos chilenos sin puntos ni comas. Ejemplo: 24900',
    )
    ws.add_data_validation(dv_precio)
    dv_precio.sqref = f'{col_precio}{DATA_START_ROW}:{col_precio}{DATA_END_ROW}'

    # ── Validación: Días de fabricación (1–60) ────────────────────────────────
    dv_dias = DataValidation(
        type='whole',
        operator='between',
        formula1='1',
        formula2='60',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='Valor inválido',
        error='Los días de fabricación deben estar entre 1 y 60',
        showInputMessage=True,
        promptTitle='Días de fabricación',
        prompt='Número de días estimados para fabricar el producto (1–60)',
    )
    ws.add_data_validation(dv_dias)
    dv_dias.sqref = f'{col_dias}{DATA_START_ROW}:{col_dias}{DATA_END_ROW}'


def _build_reference_sheet(ws2, categories, materials):
    """Construye la hoja de Referencias con listas de valores válidos."""
    ref_title_font  = Font(name='Calibri', bold=True, color='FF5C00', size=13)
    ref_header_font = Font(name='Calibri', bold=True, color='FF5C00', size=11)
    ref_value_font  = Font(name='Calibri', color='E0E0E0', size=10)
    ref_fill        = PatternFill('solid', fgColor='111214')
    ref_fill_value  = PatternFill('solid', fgColor='1A1C20')

    ws2['A1']      = 'VALORES VÁLIDOS — no modificar esta hoja'
    ws2['A1'].font = ref_title_font
    ws2['A1'].fill = ref_fill

    sections = [
        ('A', 'CATEGORÍAS DISPONIBLES', [c.name for c in categories]),
        ('C', 'MATERIALES DISPONIBLES', [m.name for m in materials]),
        ('E', 'ESTADO',                 ['activo', 'inactivo', 'pedido']),
        ('G', 'DESTACADO',              ['si', 'no']),
    ]

    for col_letter, title, values in sections:
        cell      = ws2[f'{col_letter}3']
        cell.value = title
        cell.font  = ref_header_font
        cell.fill  = ref_fill
        for i, val in enumerate(values, start=4):
            c       = ws2.cell(row=i, column=ord(col_letter) - ord('A') + 1, value=val)
            c.font  = ref_value_font
            c.fill  = ref_fill_value

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws2.column_dimensions[col].width = 24

    ws2.sheet_state = 'visible'


# ── IMPORTAR EXCEL ────────────────────────────────────────────────────────────

STATUS_MAP = {
    'activo': 'active', 'active': 'active',
    'inactivo': 'inactive', 'inactive': 'inactive',
    'pedido': 'on_request', 'on_request': 'on_request',
}


def import_products(file_obj):
    """
    Lee el Excel y crea/actualiza productos.
    Retorna (creados, actualizados, errores[])
    """
    wb      = openpyxl.load_workbook(file_obj, data_only=True)
    # Intentar leer desde la hoja "Productos", si no existe usa la activa
    ws      = wb['Productos'] if 'Productos' in wb.sheetnames else wb.active
    created = updated = 0
    errors  = []

    cats = {c.name.lower().strip(): c for c in Category.objects.all()}
    mats = {m.name.lower().strip(): m for m in Material.objects.all()}

    for row_num, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        if not row or not row[0] or str(row[0]).strip() == '':
            continue

        try:
            nombre      = str(row[0]).strip()
            cat_name    = str(row[1]).strip() if row[1] else ''
            precio_raw  = row[2]
            desc_corta  = str(row[3]).strip() if row[3] else ''
            descripcion = str(row[4]).strip() if row[4] else ''
            mats_raw    = str(row[5]).strip() if row[5] else ''
            vehiculos   = str(row[6]).strip() if row[6] else ''
            dias_fab    = int(row[7]) if row[7] else 7
            destacado   = str(row[8]).strip().lower() if row[8] else 'no'
            estado_raw  = str(row[9]).strip().lower() if row[9] else 'activo'

            if not nombre:
                errors.append(f'Fila {row_num}: nombre vacío, se omite.')
                continue

            try:
                precio = Decimal(str(precio_raw).replace(',', '').replace('.', ''))
                if precio <= 0:
                    raise ValueError
            except (InvalidOperation, ValueError, TypeError):
                errors.append(f'Fila {row_num} ({nombre}): precio inválido "{precio_raw}".')
                continue

            categoria = cats.get(cat_name.lower())
            if not categoria and cat_name:
                from django.utils.text import slugify
                categoria, _ = Category.objects.get_or_create(
                    name=cat_name,
                    defaults={'slug': slugify(cat_name)}
                )
                cats[cat_name.lower()] = categoria

            estado    = STATUS_MAP.get(estado_raw, 'active')

            from django.utils.text import slugify
            slug      = slugify(nombre)
            base_slug = slug
            counter   = 1
            while Product.objects.filter(slug=slug).exclude(name=nombre).exists():
                slug    = f'{base_slug}-{counter}'
                counter += 1

            product, is_new = Product.objects.update_or_create(
                name=nombre,
                defaults={
                    'slug':               slug,
                    'category':           categoria,
                    'base_price':         precio,
                    'short_description':  desc_corta[:300],
                    'description':        descripcion or desc_corta or nombre,
                    'compatible_vehicles': vehiculos,
                    'production_days':    dias_fab,
                    'is_featured':        destacado in ('si', 'sí', 'yes', '1', 'true'),
                    'status':             estado,
                }
            )

            if mats_raw:
                product.materials.clear()
                for mat_name in mats_raw.split('/'):
                    mat_name = mat_name.strip()
                    mat      = mats.get(mat_name.lower())
                    if mat:
                        product.materials.add(mat)
                    elif mat_name:
                        new_mat, _ = Material.objects.get_or_create(
                            name=mat_name,
                            defaults={'material_type': 'polyurethane'}
                        )
                        product.materials.add(new_mat)
                        mats[mat_name.lower()] = new_mat

            if is_new:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors.append(f'Fila {row_num}: error inesperado — {str(e)}')

    return created, updated, errors
