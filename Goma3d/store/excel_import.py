"""
Importación y exportación de productos desde/hacia Excel para Gflex3D.

Columnas de la plantilla:
  nombre | categoria | precio | descripcion_corta | descripcion |
  materiales | vehiculos_compatibles | dias_fabricacion | stock |
  precio_minimo | destacado | estado
"""
import io
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from .models import Product, Category, Material, ProductVariant


# ── COLUMNAS ─────────────────────────────────────────────────────────────────

COLUMNS = [
    ('nombre',               'Nombre del producto *',          28),
    ('categoria',            'Categoría *',                    18),
    ('precio',               'Precio base CLP *',              14),
    ('descripcion_corta',    'Descripción corta',              30),
    ('descripcion',          'Descripción completa',           40),
    ('materiales',           'Materiales (separados por /)',   24),
    ('vehiculos',            'Vehículos compatibles',          30),
    ('dias_fabricacion',     'Días de fabricación',            14),
    ('destacado',            'Destacado (si/no)',              14),
    ('estado',               'Estado (activo/inactivo/pedido)',18),
]

HEADERS    = [c[1] for c in COLUMNS]
COL_WIDTHS = [c[2] for c in COLUMNS]
COL_KEYS   = [c[0] for c in COLUMNS]


# ── GENERAR PLANTILLA ─────────────────────────────────────────────────────────

def generate_template(categories, materials):
    """Genera el archivo Excel plantilla para descargar."""
    wb = openpyxl.Workbook()

    # ── Hoja principal ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Productos'

    # Estilo encabezados
    header_fill   = PatternFill('solid', fgColor='0A0A0B')
    header_font   = Font(name='Calibri', bold=True, color='FF5C00', size=11)
    subhead_fill  = PatternFill('solid', fgColor='1A1C20')
    subhead_font  = Font(name='Calibri', color='F5F3EF', size=10)
    border_side   = Side(style='thin', color='333333')
    cell_border   = Border(left=border_side, right=border_side,
                           bottom=border_side, top=border_side)

    # Fila 1 — título
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value          = '🔩 GFLEX3D — Importación de productos'
    title_cell.font           = Font(name='Calibri', bold=True, color='FF5C00', size=14)
    title_cell.fill           = PatternFill('solid', fgColor='0A0A0B')
    title_cell.alignment      = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Fila 2 — instrucción
    ws.merge_cells('A2:J2')
    inst = ws['A2']
    inst.value     = 'Completa desde la fila 4. Columnas con * son obligatorias. No modifiques los encabezados.'
    inst.font      = Font(name='Calibri', color='8A8F9A', size=10, italic=True)
    inst.fill      = PatternFill('solid', fgColor='111214')
    inst.alignment = Alignment(horizontal='center')

    # Fila 3 — encabezados
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell            = ws.cell(row=3, column=col_idx, value=header)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border     = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 32

    # Filas de ejemplo (4–6)
    examples = [
        ['Buje Estabilizador Delantero x2', 'Bujes', 24900,
         'Buje de poliuretano para estabilizador',
         'Fabricado en poliuretano de alta durabilidad. Reemplaza el buje original con mayor vida útil.',
         'Poliuretano 80A', 'Toyota Corolla\nNissan Sentra\nHyundai Accent',
         7, 'si', 'activo'],
        ['Manguera Radiador Superior Universal', 'Mangueras', 12900,
         'Manguera de silicona para radiador',
         'Compatible con vehículos nacionales e importados. Alta resistencia al calor.',
         'Caucho EPDM', 'Toyota\nNissan\nHyundai\nKia',
         5, 'no', 'activo'],
        ['Guardapolvos Amortiguador Trasero', 'Guardapolvos', 8500,
         'Guardapolvos en caucho NBR',
         'Protege el amortiguador del polvo y la humedad. Fácil instalación.',
         'Caucho NBR', 'Suzuki Swift\nSuzuki Baleno',
         7, 'no', 'pedido'],
    ]

    example_fill = PatternFill('solid', fgColor='1A1C20')
    example_font = Font(name='Calibri', color='C5C7CC', size=10)

    for row_idx, row_data in enumerate(examples, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell            = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font       = example_font
            cell.fill       = example_fill
            cell.border     = cell_border
            cell.alignment  = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[row_idx].height = 40

    # Filas vacías para rellenar (7–56)
    empty_fill = PatternFill('solid', fgColor='111214')
    empty_font = Font(name='Calibri', color='555960', size=10)
    for row_idx in range(7, 57):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell        = ws.cell(row=row_idx, column=col_idx, value='')
            cell.font   = empty_font
            cell.fill   = empty_fill
            cell.border = cell_border
        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = 'A4'

    # ── Hoja de referencia ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Referencias')
    ws2['A1'] = '📋 VALORES VÁLIDOS'
    ws2['A1'].font = Font(bold=True, color='FF5C00', size=13)

    ws2['A3'] = 'CATEGORÍAS DISPONIBLES'
    ws2['A3'].font = Font(bold=True, color='FF5C00')
    for i, cat in enumerate(categories, start=4):
        ws2.cell(row=i, column=1, value=cat.name)

    col_offset = 3
    ws2.cell(row=3, column=col_offset, value='MATERIALES DISPONIBLES')
    ws2.cell(row=3, column=col_offset).font = Font(bold=True, color='FF5C00')
    for i, mat in enumerate(materials, start=4):
        ws2.cell(row=i, column=col_offset, value=mat.name)

    col_offset = 5
    ws2.cell(row=3, column=col_offset, value='ESTADO')
    ws2.cell(row=3, column=col_offset).font = Font(bold=True, color='FF5C00')
    for i, val in enumerate(['activo', 'inactivo', 'pedido'], start=4):
        ws2.cell(row=i, column=col_offset, value=val)

    col_offset = 7
    ws2.cell(row=3, column=col_offset, value='DESTACADO')
    ws2.cell(row=3, column=col_offset).font = Font(bold=True, color='FF5C00')
    for i, val in enumerate(['si', 'no'], start=4):
        ws2.cell(row=i, column=col_offset, value=val)

    for col in ['A','B','C','D','E','F','G']:
        ws2.column_dimensions[col].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── IMPORTAR EXCEL ────────────────────────────────────────────────────────────

STATUS_MAP = {
    'activo': 'active', 'active': 'active',
    'inactivo': 'inactive', 'inactive': 'inactive',
    'pedido': 'on_request', 'on_request': 'on_request',
}


def import_products(file_obj):
    """
    Lee el Excel y crea/actualiza productos.
    Retorna (creados, actualizados, errores[])
    """
    wb     = openpyxl.load_workbook(file_obj, data_only=True)
    ws     = wb.active
    created = updated = 0
    errors  = []

    # Cargar categorías y materiales en dicts para lookup rápido
    cats = {c.name.lower().strip(): c for c in Category.objects.all()}
    mats = {m.name.lower().strip(): m for m in Material.objects.all()}

    # Leer desde fila 4 (3 son títulos/ejemplos)
    for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Saltar filas vacías
        if not row or not row[0] or str(row[0]).strip() == '':
            continue

        try:
            nombre       = str(row[0]).strip()
            cat_name     = str(row[1]).strip() if row[1] else ''
            precio_raw   = row[2]
            desc_corta   = str(row[3]).strip() if row[3] else ''
            descripcion  = str(row[4]).strip() if row[4] else ''
            mats_raw     = str(row[5]).strip() if row[5] else ''
            vehiculos    = str(row[6]).strip() if row[6] else ''
            dias_fab     = int(row[7]) if row[7] else 7
            destacado    = str(row[8]).strip().lower() if row[8] else 'no'
            estado_raw   = str(row[9]).strip().lower() if row[9] else 'activo'

            # Validar nombre
            if not nombre:
                errors.append(f'Fila {row_num}: nombre vacío, se omite.')
                continue

            # Validar precio
            try:
                precio = Decimal(str(precio_raw).replace(',', '').replace('.', ''))
                if precio <= 0:
                    raise ValueError
            except (InvalidOperation, ValueError, TypeError):
                errors.append(f'Fila {row_num} ({nombre}): precio inválido "{precio_raw}".')
                continue

            # Categoría
            categoria = cats.get(cat_name.lower())
            if not categoria and cat_name:
                # Crear categoría si no existe
                from django.utils.text import slugify
                categoria, _ = Category.objects.get_or_create(
                    name=cat_name,
                    defaults={'slug': slugify(cat_name)}
                )
                cats[cat_name.lower()] = categoria

            # Estado
            estado = STATUS_MAP.get(estado_raw, 'active')

            # Crear o actualizar producto
            from django.utils.text import slugify
            slug = slugify(nombre)
            # Si hay slug duplicado, agregar sufijo
            base_slug = slug
            counter   = 1
            while Product.objects.filter(slug=slug).exclude(name=nombre).exists():
                slug = f'{base_slug}-{counter}'
                counter += 1

            product, is_new = Product.objects.update_or_create(
                name=nombre,
                defaults={
                    'slug':              slug,
                    'category':          categoria,
                    'base_price':        precio,
                    'short_description': desc_corta[:300],
                    'description':       descripcion or desc_corta or nombre,
                    'compatible_vehicles': vehiculos,
                    'production_days':   dias_fab,
                    'is_featured':       destacado in ('si', 'sí', 'yes', '1', 'true'),
                    'status':            estado,
                }
            )

            # Asignar materiales
            if mats_raw:
                product.materials.clear()
                for mat_name in mats_raw.split('/'):
                    mat_name = mat_name.strip()
                    mat = mats.get(mat_name.lower())
                    if mat:
                        product.materials.add(mat)
                    elif mat_name:
                        # Crear material si no existe
                        new_mat, _ = Material.objects.get_or_create(
                            name=mat_name,
                            defaults={'material_type': 'polyurethane'}
                        )
                        product.materials.add(new_mat)
                        mats[mat_name.lower()] = new_mat

            if is_new:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors.append(f'Fila {row_num}: error inesperado — {str(e)}')

    return created, updated, errors
